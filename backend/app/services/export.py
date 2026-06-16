import io
from datetime import datetime
from typing import List

from fastapi.responses import StreamingResponse


def export_excel(warga_list: list) -> StreamingResponse:
    """Generate Excel dari daftar warga."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Prioritas Warga"

    # header
    headers = [
        "No", "Nama KK", "Nomor KK", "Skor Kebutuhan", "Kategori",
        "Kondisi Hunian", "Total Anggota", "Air Bersih", "Listrik",
        "Status Toilet", "Status Rumah", "Rata Sekolah (thn)", "Tanggal Input"
    ]
    header_fill = PatternFill("solid", fgColor="1B1B1B")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # data rows
    for i, w in enumerate(warga_list, 1):
        ws.append([
            i,
            w.get("nama", ""),
            w.get("nomorKK", ""),
            w.get("score", 0),
            w.get("kategori", {}).get("label", ""),
            w.get("kondisi", {}).get("label", ""),
            w.get("totalAnggota", ""),
            "Ya" if w.get("airBersih") else "Tidak",
            "Ya" if w.get("adaListrik") else "Tidak",
            w.get("statusToilet", ""),
            w.get("statusRumah", ""),
            w.get("rataSekolah", ""),
            w.get("tanggal", ""),
        ])

    # auto-width kolom
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"siprios_prioritas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def export_pdf(warga_list: list) -> StreamingResponse:
    """Generate PDF dari daftar warga."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    # judul
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=14, spaceAfter=8)
    elements.append(Paragraph("Daftar Prioritas Penerima Bantuan Sosial", title_style))
    elements.append(Paragraph(
        f"Dicetak: {datetime.now().strftime('%d %B %Y, %H:%M')} · SIPRIOS",
        styles["Normal"]
    ))
    elements.append(Spacer(1, 0.4*cm))

    # tabel
    table_data = [["No", "Nama KK", "Nomor KK", "Skor", "Kategori", "Kondisi Hunian", "Anggota", "Tanggal"]]
    for i, w in enumerate(warga_list, 1):
        tgl = w.get("tanggal", "")
        if hasattr(tgl, "strftime"):
            tgl = tgl.strftime("%d/%m/%Y")
        table_data.append([
            str(i),
            w.get("nama", "")[:25],
            w.get("nomorKK", ""),
            str(w.get("score", 0)),
            w.get("kategori", {}).get("label", ""),
            w.get("kondisi", {}).get("label", "")[:20],
            str(w.get("totalAnggota", "")),
            str(tgl)[:10],
        ])

    col_widths = [1*cm, 5*cm, 4*cm, 1.5*cm, 3*cm, 5*cm, 2*cm, 2.5*cm]
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B1B1B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (3, 0), (3, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D3D3D3")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(tbl)

    doc.build(elements)
    buf.seek(0)

    filename = f"siprios_prioritas_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
